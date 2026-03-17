import json
from datetime import date, datetime
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook

from ..extensions import db
from ..models import (
    FinalTableConfig,
    SettlementFundBatch,
    SettlementFundRecord,
    SourceTableConfig,
    StrategyInsuranceMap,
    StrategySectorMap,
)
from ..permissions import permission_required


dashboard_bp = Blueprint("dashboard", __name__)


@dashboard_bp.route("/")
@login_required
def index():
    return redirect(url_for("dashboard.final_tables"))


@dashboard_bp.route("/data-sources")
@login_required
@permission_required("data_sources.view")
def data_sources():
    return redirect(url_for("dashboard.data_source_settlement_fund"))


@dashboard_bp.route("/data-sources/settlement-fund")
@login_required
@permission_required("data_sources.view")
def data_source_settlement_fund():
    selected_date = _parse_input_date(request.args.get("settlement_date", "").strip())
    strategy_code_filter = request.args.get("strategy_code", "").strip()
    strategy_code_keyword = request.args.get("strategy_code_keyword", "").strip()
    sector_filter = request.args.get("sector", "").strip()
    insurance_type_filter = request.args.get("insurance_type", "").strip()
    if not selected_date:
        latest_date = (
            db.session.query(SettlementFundRecord.settlement_date)
            .filter(SettlementFundRecord.settlement_date.isnot(None))
            .order_by(SettlementFundRecord.settlement_date.desc())
            .first()
        )
        selected_date = latest_date[0] if latest_date else None

    query = SettlementFundRecord.query
    if selected_date:
        query = query.filter(SettlementFundRecord.settlement_date == selected_date)
    records = query.order_by(SettlementFundRecord.id.asc()).all()

    strategy_codes = []
    for record in records:
        strategy_code = _to_text(record.row_data.get("策略账户"))
        if strategy_code:
            strategy_codes.append(strategy_code)
    unique_strategy_codes = sorted(set(strategy_codes))

    sector_map = {}
    insurance_map = {}
    if unique_strategy_codes:
        sector_map = {
            x.strategy: x.sector
            for x in StrategySectorMap.query.filter(
                StrategySectorMap.strategy.in_(unique_strategy_codes)
            ).all()
        }
        insurance_map = {
            x.strategy: x.insurance_type
            for x in StrategyInsuranceMap.query.filter(
                StrategyInsuranceMap.strategy.in_(unique_strategy_codes)
            ).all()
        }

    display_columns = [
        "结算日期",
        "策略代码",
        "板块",
        "策略属性",
        "手续费",
        "交割手续费",
        "行权手续费",
        "申报费",
        "递延费",
        "保证金",
        "多头保证金",
        "空头保证金",
        "当日逐日盈亏总和",
    ]

    display_rows = []
    for record in records:
        data = record.row_data or {}
        strategy_code = _to_text(data.get("策略账户"))
        display_rows.append(
            {
                "结算日期": _format_date_for_view(record.settlement_date or data.get("结算日期")),
                "策略代码": strategy_code or "-",
                "板块": sector_map.get(strategy_code, "-"),
                "策略属性": insurance_map.get(strategy_code, "-"),
                "手续费": _pick_field(data, "手续费"),
                "交割手续费": _pick_field(data, "交割手续费"),
                "行权手续费": _pick_field(data, "行权手续费"),
                "申报费": _pick_field(data, "申报费"),
                "递延费": _pick_field(data, "递延费"),
                "保证金": _pick_field(data, "保证金", "总保证金"),
                "多头保证金": _pick_field(data, "多头保证金"),
                "空头保证金": _pick_field(data, "空头保证金"),
                "当日逐日盈亏总和": _pick_field(
                    data, "当日逐日盈亏总和（含期权）", "当日逐日盈亏总和"
                ),
            }
        )

    strategy_code_options = sorted(
        {row["策略代码"] for row in display_rows if row["策略代码"] not in ("", "-")}
    )
    sector_options = sorted({row["板块"] for row in display_rows if row["板块"] not in ("", "-")})
    insurance_type_options = sorted(
        {row["策略属性"] for row in display_rows if row["策略属性"] not in ("", "-")}
    )

    filtered_rows = display_rows
    if strategy_code_filter:
        filtered_rows = [r for r in filtered_rows if r["策略代码"] == strategy_code_filter]
    if strategy_code_keyword:
        kw = strategy_code_keyword.lower()
        filtered_rows = [r for r in filtered_rows if kw in _to_text(r["策略代码"]).lower()]
    if sector_filter:
        filtered_rows = [r for r in filtered_rows if r["板块"] == sector_filter]
    if insurance_type_filter:
        filtered_rows = [r for r in filtered_rows if r["策略属性"] == insurance_type_filter]

    raw_available_dates = [
        r[0]
        for r in db.session.query(SettlementFundRecord.settlement_date)
        .filter(SettlementFundRecord.settlement_date.isnot(None))
        .distinct()
        .order_by(SettlementFundRecord.settlement_date.desc())
        .all()
    ]
    available_dates = []
    for raw in raw_available_dates:
        parsed = _normalize_settlement_date(raw)
        if parsed:
            date_str = parsed.strftime("%Y-%m-%d")
            available_dates.append({"date": parsed, "date_str": date_str})

    selected_date_str = selected_date.strftime("%Y-%m-%d") if selected_date else ""
    return render_template(
        "dashboard/data_source_settlement_fund.html",
        display_columns=display_columns,
        display_rows=filtered_rows,
        strategy_code_options=strategy_code_options,
        sector_options=sector_options,
        insurance_type_options=insurance_type_options,
        strategy_code_filter=strategy_code_filter,
        strategy_code_keyword=strategy_code_keyword,
        sector_filter=sector_filter,
        insurance_type_filter=insurance_type_filter,
        available_dates=available_dates,
        selected_date=selected_date,
        selected_date_str=selected_date_str,
    )


@dashboard_bp.route("/data-sources/settlement-position-summary")
@login_required
@permission_required("data_sources.view")
def data_source_settlement_position_summary():
    items = SourceTableConfig.query.order_by(SourceTableConfig.id.desc()).all()
    return render_template("dashboard/data_source_settlement_position_summary.html", items=items)


@dashboard_bp.route("/final-tables")
@login_required
@permission_required("final_tables.view")
def final_tables():
    items = FinalTableConfig.query.order_by(FinalTableConfig.id.desc()).all()
    return render_template("dashboard/final_tables.html", items=items)


@dashboard_bp.route("/data-sources/settlement-fund", methods=["POST"])
@login_required
@permission_required("data_sources.view")
def import_data_source_settlement_fund():
    return _import_settlement_fund_excel()


def _import_settlement_fund_excel():
    excel_file = request.files.get("excel_file")
    if not excel_file or not excel_file.filename:
        flash("请先选择要上传的 Excel 文件", "error")
        return redirect(url_for("dashboard.data_source_settlement_fund"))

    if not excel_file.filename.lower().endswith(".xlsx"):
        flash("当前仅支持 .xlsx 文件", "error")
        return redirect(url_for("dashboard.data_source_settlement_fund"))

    try:
        file_bytes = excel_file.read()
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        headers, rows = _extract_excel_rows(sheet)
    except Exception:
        flash("Excel 解析失败，请确认文件格式正确", "error")
        return redirect(url_for("dashboard.data_source_settlement_fund"))

    if not headers or not rows:
        flash("未识别到有效数据，请检查 Excel 结构", "error")
        return redirect(url_for("dashboard.data_source_settlement_fund"))

    batch = SettlementFundBatch(
        file_name=excel_file.filename,
        sheet_name=sheet.title,
        headers_json=json.dumps(headers, ensure_ascii=False),
        row_count=len(rows),
        uploaded_by_id=current_user.id,
    )
    db.session.add(batch)
    db.session.flush()

    for row_no, row_data in enumerate(rows, start=1):
        db.session.add(
            SettlementFundRecord(
                batch_id=batch.id,
                row_no=row_no,
                settlement_date=_normalize_settlement_date(row_data.get("结算日期")),
                row_data=row_data,
            )
        )

    db.session.commit()
    flash(f"导入成功，共写入 {len(rows)} 条数据", "success")
    first_row_date = _normalize_settlement_date(rows[0].get("结算日期")) if rows else None
    if first_row_date:
        return redirect(
            url_for(
                "dashboard.data_source_settlement_fund",
                settlement_date=first_row_date.strftime("%Y-%m-%d"),
            )
        )
    return redirect(url_for("dashboard.data_source_settlement_fund"))


def _extract_excel_rows(sheet):
    header_row = None
    header_row_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [str(v).strip() if v is not None else "" for v in row]
        if "序号" in values:
            header_row = values
            header_row_idx = row_idx
            break

    if not header_row or not header_row_idx:
        return [], []

    active_indexes = [idx for idx, val in enumerate(header_row) if val]
    headers = [header_row[idx] for idx in active_indexes]

    rows = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        values = [row[idx] if idx < len(row) else None for idx in active_indexes]
        if all(v in (None, "") for v in values):
            continue
        row_data = {}
        for i, key in enumerate(headers):
            value = values[i]
            if isinstance(value, str):
                row_data[key] = value.strip()
            else:
                row_data[key] = value
        rows.append(row_data)

    return headers, rows


def _normalize_settlement_date(raw_value):
    if raw_value is None:
        return None

    if isinstance(raw_value, date):
        return raw_value if not isinstance(raw_value, datetime) else raw_value.date()

    value = str(raw_value).strip()
    if not value:
        return None

    value = value.replace("-", "").replace("/", "")
    if value.endswith(".0"):
        value = value[:-2]
    if len(value) == 8 and value.isdigit():
        try:
            return datetime.strptime(value, "%Y%m%d").date()
        except ValueError:
            return None
    return None


def _parse_input_date(raw_value: str):
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _pick_field(data, *keys):
    for key in keys:
        value = data.get(key)
        if value not in (None, ""):
            return value
    return "-"


def _format_date_for_view(value):
    parsed = _normalize_settlement_date(value)
    if parsed:
        return parsed.strftime("%Y-%m-%d")
    return _to_text(value) or "-"
